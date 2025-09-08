import asyncio
import contextlib
import datetime as dt
import json
import os
import signal
import sys
import csv
from dataclasses import dataclass, asdict
from typing import List, Optional

from dateutil import tz
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from playwright.async_api import async_playwright, Page


STREAMYARD_URL_DEFAULT = "https://streamyard.studio/?v=UnchainedPodcasts"


@dataclass
class ChatMessage:
    message: str
    nickname: str
    start_time: str
    end_time: str
    message_time: Optional[str] = ""


def _now_iso_local() -> str:
    now = dt.datetime.now(tz.tzlocal())
    return now.replace(microsecond=0).isoformat()


async def _ensure_playwright_browsers_installed() -> None:
    # Best-effort install if missing
    try:
        from playwright.__main__ import main as playwright_main  # type: ignore

        # Equivalent to: playwright install chromium
        sys.argv = ["playwright", "install", "chromium", "--with-deps"]
        playwright_main()
    except Exception:
        # If install already present or offline, proceed
        pass


async def export_to_excel(records: List[ChatMessage], output_path: str, include_message_time: bool = False) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Messages"

    headers = ["Message", "Nickname", "Start Time", "End Time"]
    if include_message_time:
        headers.append("Message Time")
    ws.append(headers)
    for record in records:
        row = [record.message, record.nickname, record.start_time, record.end_time]
        if include_message_time:
            row.append(record.message_time or "")
        ws.append(row)

    # Autofit basic column widths
    for idx, header in enumerate(headers, start=1):
        max_len = max(
            [len(str(header))]
            + [len(str(ws.cell(row=row, column=idx).value or "")) for row in range(2, ws.max_row + 1)]
        )
        ws.column_dimensions[get_column_letter(idx)].width = min(60, max(12, max_len + 2))

    os.makedirs(os.path.dirname(os.path.abspath(output_path)) or ".", exist_ok=True)
    wb.save(output_path)


async def export_to_csv(records: List[ChatMessage], output_path: str, include_message_time: bool = False) -> None:
    os.makedirs(os.path.dirname(os.path.abspath(output_path)) or ".", exist_ok=True)
    headers = ["Message", "Nickname", "Start Time", "End Time"]
    if include_message_time:
        headers.append("Message Time")
    with open(output_path, "w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        writer.writerow(headers)
        for record in records:
            row = [record.message, record.nickname, record.start_time, record.end_time]
            if include_message_time:
                row.append(record.message_time or "")
            writer.writerow(row)


async def collect_streamyard_chat(
    url: str,
    output_path: str,
    css_selector_message_container: Optional[str] = None,
    css_selector_each_message: Optional[str] = None,
    css_selector_nickname: Optional[str] = None,
    css_selector_text: Optional[str] = None,
    include_message_time: bool = False,
    also_write_csv: bool = False,
) -> List[ChatMessage]:
    """
    Opens the given StreamYard studio URL, waits for user to log in if needed,
    listens for chat updates, and records messages until the user stops via Ctrl+C.

    If custom CSS selectors are provided, uses them; otherwise attempts StreamYard defaults.
    """

    start_time = _now_iso_local()
    messages: List[ChatMessage] = []

    await _ensure_playwright_browsers_installed()

    async with async_playwright() as p:
        browser = await p.chromium.launch(headless=False)
        context = await browser.new_context()
        page = await context.new_page()

        await page.goto(url, wait_until="domcontentloaded")

        # Allow manual login/2FA; wait until chat panel appears or timeout
        await _wait_for_chat_ready(page)

        # Determine selectors
        sel_container, sel_each, sel_nick, sel_text = _resolve_selectors(
            css_selector_message_container,
            css_selector_each_message,
            css_selector_nickname,
            css_selector_text,
        )

        seen_ids = set()

        print("Collecting messages. Press Ctrl+C to stop and export...")
        try:
            while True:
                new_items = await _extract_messages(page, sel_container, sel_each, sel_nick, sel_text)
                for m_id, nickname, text in new_items:
                    if m_id in seen_ids:
                        continue
                    seen_ids.add(m_id)
                    messages.append(
                        ChatMessage(
                            message=text,
                            nickname=nickname,
                            start_time=start_time,
                            end_time="",
                            message_time=_now_iso_local() if include_message_time else "",
                        )
                    )
                await asyncio.sleep(1.0)
        except KeyboardInterrupt:
            pass
        finally:
            end_time = _now_iso_local()
            for msg in messages:
                msg.end_time = end_time
            await context.close()
            await browser.close()

    await export_to_excel(messages, output_path, include_message_time=include_message_time)
    if also_write_csv:
        root, _ = os.path.splitext(output_path)
        csv_path = root + ".csv"
        await export_to_csv(messages, csv_path, include_message_time=include_message_time)
    return messages


async def _wait_for_chat_ready(page: Page) -> None:
    # Try a few likely selectors, fall back to timeout
    candidate_selectors = [
        # StreamYard chat area candidates (may change over time)
        "[data-testid=chat-pane]",
        "[class*='Chat']",
        "[role=log]",
    ]
    for selector in candidate_selectors:
        try:
            await page.wait_for_selector(selector, timeout=15000)
            return
        except Exception:
            continue
    # If none matched, just continue; user can provide custom selectors
    return


def _resolve_selectors(
    css_selector_message_container: Optional[str],
    css_selector_each_message: Optional[str],
    css_selector_nickname: Optional[str],
    css_selector_text: Optional[str],
):
    # Defaults tailored for StreamYard as of 2025-09; may need updates
    default_container = "[data-testid=chat-pane]"
    default_each = "[data-testid=chat-message]"
    default_nick = "[data-testid=chat-message-author]"
    default_text = "[data-testid=chat-message-content]"

    return (
        css_selector_message_container or default_container,
        css_selector_each_message or default_each,
        css_selector_nickname or default_nick,
        css_selector_text or default_text,
    )


async def _extract_messages(
    page: Page,
    sel_container: str,
    sel_each: str,
    sel_nick: str,
    sel_text: str,
):
    # Returns list of tuples: (unique_id, nickname, text)
    js = f"""
    (selContainer, selEach, selNick, selText) => {{
      const container = document.querySelector(selContainer) || document;
      const nodes = Array.from(container.querySelectorAll(selEach));
      return nodes.map((node, idx) => {{
        const nickNode = node.querySelector(selNick);
        const textNode = node.querySelector(selText);
        const nickname = nickNode ? nickNode.textContent?.trim() || '' : '';
        const text = textNode ? textNode.textContent?.trim() || '' : '';
        const uid = node.getAttribute('data-message-id') || node.id || `${{Date.now()}}-${{idx}}-${{nickname}}-${{text}}`;
        return [uid, nickname, text];
      }});
    }}
    """
    results = await page.evaluate(js, sel_container, sel_each, sel_nick, sel_text)
    # Filter empties
    cleaned = []
    for item in results:
        try:
            uid, nickname, text = item
        except Exception:
            continue
        if not text and not nickname:
            continue
        cleaned.append((str(uid), str(nickname), str(text)))
    return cleaned


async def main() -> None:
    import argparse

    parser = argparse.ArgumentParser(description="StreamYard chat scraper to Excel (.xlsx)")
    parser.add_argument("--url", default=STREAMYARD_URL_DEFAULT, help="StreamYard studio URL")
    parser.add_argument(
        "--output",
        default="output/streamyard_chat.xlsx",
        help="Path to save Excel file",
    )
    parser.add_argument("--container", help="CSS selector: chat container", default=None)
    parser.add_argument("--each", help="CSS selector: each message", default=None)
    parser.add_argument("--nick", help="CSS selector: nickname", default=None)
    parser.add_argument("--text", help="CSS selector: message text", default=None)
    parser.add_argument("--with-message-time", action="store_true", help="Include per-message timestamp column")
    parser.add_argument("--csv", action="store_true", help="Also write CSV next to the Excel file")

    args = parser.parse_args()

    await collect_streamyard_chat(
        url=args.url,
        output_path=args.output,
        css_selector_message_container=args.container,
        css_selector_each_message=args.each,
        css_selector_nickname=args.nick,
        css_selector_text=args.text,
        include_message_time=bool(args.with_message_time),
        also_write_csv=bool(args.csv),
    )


if __name__ == "__main__":
    try:
        asyncio.run(main())
    except KeyboardInterrupt:
        pass

